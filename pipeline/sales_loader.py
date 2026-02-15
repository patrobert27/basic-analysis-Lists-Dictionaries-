from typing import Any
from openpyxl import load_workbook


def _sheet_to_dicts(path: str, sheet_name: str) -> list[dict[str, Any]]:
    """
    Internal helper.
    Reads a sheet from an Excel file and converts it into list[dict].
    """
    rows: list[dict[str, Any]] = []

    try:
        wb = load_workbook(path, data_only=True)
        sheet = wb[sheet_name]

        # Read all rows
        data = list(sheet.iter_rows(values_only=True))

        if not data:
            return []

        headers = data[0]

        for row_values in data[1:]:
            row_dict = dict(zip(headers, row_values))
            rows.append(row_dict)

        return rows

    except KeyError:
        print(f"Sheet '{sheet_name}' not found.")
        return []
    except OSError as e:
        print(f"File error: {e}")
        return []


def load_transactions(path: str) -> list[dict[str, Any]]:
    """
    Load Transactions sheet as list[dict].
    """
    return _sheet_to_dicts(path, "Transactions")


def load_targets(path: str) -> list[dict[str, Any]]:
    """
    Load Targets sheet as list[dict].
    """
    return _sheet_to_dicts(path, "Targets")
